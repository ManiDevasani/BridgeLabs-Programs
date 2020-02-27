from HashingList import *


class LaunchInstance:
    if __name__ == '__main__':
        import openpyxl
        import boto3
        ec2 = boto3.resource('ec2')
        tag_purpose = {"Key": "Name", "Value": "Trails"}
        path = "/home/admin2/Downloads/Manikanta/ec2-instance.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        # calling the main method and creating the object of the hashing
        HashMap = HashingList()
        rows = sheet.max_row + 1
        column = sheet.max_column + 1
        for i in range(1, rows):
            for j in range(1, column):
                if i is not 1:
                    value = sheet.cell(row=i, column=j).value
                    HashMap.insert(value, j - 1)
        for i in range(1, rows - 1):
            conform = HashMap.Conformation(i)
            Storage_size = HashMap.storage(i)
            if conform is None:
                instances = ec2.create_instances(
                    InstanceType=HashMap.instance_type(i),
                    ImageId=HashMap.Image_id(i),
                    MinCount=HashMap.min_count(i),
                    MaxCount=HashMap.max_count(i),
                    BlockDeviceMappings=[{"DeviceName": "/dev/sdh", "Ebs": {"VolumeSize": Storage_size}}],
                    KeyName="Trail",
                    TagSpecifications=[{'ResourceType': 'instance', 'Tags': [tag_purpose]}]
                )
                c2 = sheet.cell(row=i + 1, column=7)
                c2.value = "ok"
                wb.save(path)
